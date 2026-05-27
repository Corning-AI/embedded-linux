"""
Generate publication-quality charts for V2 fresh-ice NIRS experiment.

Charts produced (PNG, saved to cold_therapy_showcase/):
  13_v2_dose_response_peak.png       — Cold peak (S1) vs ice duration
  14_v2_dose_response_residual.png   — Sample-5 residual vs ice duration
  15_v2_trajectories.png             — All 5 within-window trajectories
  16_v2_nir_vs_pt100.png             — NIRS dose-response overlaid on PT100 cooling curve
"""
import os
import csv
from pathlib import Path
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

plt.rcParams.update({
    "font.size": 12,
    "axes.titlesize": 14,
    "axes.labelsize": 12,
    "legend.fontsize": 10,
    "figure.facecolor": "#0E0E14",
    "axes.facecolor": "#0E0E14",
    "axes.edgecolor": "#888888",
    "axes.labelcolor": "#DDDDDD",
    "xtick.color": "#DDDDDD",
    "ytick.color": "#DDDDDD",
    "text.color": "#FFFFFF",
    "axes.titlecolor": "#FFFFFF",
    "grid.color": "#333344",
    "savefig.facecolor": "#0E0E14",
    "savefig.edgecolor": "#0E0E14",
    "savefig.dpi": 150,
})

BASE = Path(r"c:/Users/corni/OneDrive - Nanyang Technological University/ntu_rf/M266_EmbeddedLinux")
V2DIR = BASE / "nir_intensity_20260527" / "v2_fresh_ice"
OUT_DIR = BASE / "cold_therapy_showcase"
PT100_SRC = BASE / "cold_therapy_showcase" / "12_skin_cooling_8sites.xlsx"

RUNS = [
    (12, "run1v2_12s.csv", "#FF3B30"),
    (24, "run2v2_24s.csv", "#FF9500"),
    (36, "run3v2_36s.csv", "#FFCC00"),
    (48, "run4v2_48s.csv", "#5AC8FA"),
    (60, "run5v2_60s.csv", "#AF52DE"),
]


def parse(p):
    rows = []
    with open(p, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(",")
            if parts[0] == "sample":
                header = parts
                continue
            rows.append({h: parts[i] for i, h in enumerate(header)})
    return rows


def main():
    data = {}
    for ice, fn, _color in RUNS:
        data[ice] = parse(V2DIR / fn)

    ice_durations = [r[0] for r in RUNS]
    peaks = [float(data[d][0]["toi_cal"]) for d in ice_durations]
    s5s = [float(data[d][-1]["toi_cal"]) for d in ice_durations]

    # -------- Chart 13: Cold peak dose-response --------
    fig, ax = plt.subplots(figsize=(9, 6))
    ax.plot(ice_durations, peaks, marker="o", markersize=14,
            linewidth=2.5, color="#FF3B30",
            markerfacecolor="#FF3B30", markeredgecolor="#FFFFFF",
            markeredgewidth=1.5, label="Cold peak (sample 1)")
    for x, y in zip(ice_durations, peaks):
        ax.annotate(f"{y:+.3f}", (x, y), xytext=(0, 14),
                    textcoords="offset points", ha="center",
                    fontsize=11, color="#FFFFFF",
                    bbox=dict(boxstyle="round,pad=0.25", fc="#FF3B30", ec="none", alpha=0.85))
    ax.axhline(-0.21, color="#5AC8FA", linestyle=":", linewidth=1.4, alpha=0.7,
               label="Historical baseline (−0.21)")
    ax.set_xlabel("Ice Duration (seconds)")
    ax.set_ylabel("TOI_cal (less negative = stronger cold response)")
    ax.set_title("Cold Peak Dose-Response — Fresh Ice, Fingertip (V2)")
    ax.invert_yaxis()  # so "more cold" is up
    ax.set_xticks(ice_durations)
    ax.grid(True, alpha=0.25)
    ax.legend(loc="center right")
    # Annotate saturation region (placed above plateau in inverted axis = visually below)
    ax.annotate("Saturation onset 36-48s\nplateau ~ −0.16",
                xy=(42, -0.165), xytext=(30, -0.205),
                fontsize=11, color="#FFCC00",
                arrowprops=dict(arrowstyle="->", color="#FFCC00", lw=1.5),
                ha="center")
    plt.tight_layout()
    out1 = OUT_DIR / "13_v2_dose_response_peak.png"
    plt.savefig(out1)
    plt.close(fig)
    print(f"Saved: {out1}")

    # -------- Chart 14: Sample-5 residual --------
    fig, ax = plt.subplots(figsize=(9, 6))
    ax.plot(ice_durations, s5s, marker="s", markersize=14,
            linewidth=2.5, color="#34C759",
            markerfacecolor="#34C759", markeredgecolor="#FFFFFF",
            markeredgewidth=1.5, label="Sample 5 (10s residual)")
    for x, y in zip(ice_durations, s5s):
        ax.annotate(f"{y:+.3f}", (x, y), xytext=(0, 14),
                    textcoords="offset points", ha="center",
                    fontsize=11, color="#FFFFFF",
                    bbox=dict(boxstyle="round,pad=0.25", fc="#34C759", ec="none", alpha=0.85))
    ax.set_xlabel("Ice Duration (seconds)")
    ax.set_ylabel("TOI_cal at +13s after ice removal")
    ax.set_title("Sample-5 Residual Dose-Response — Fresh Ice, Fingertip (V2)")
    ax.invert_yaxis()
    ax.set_xticks(ice_durations)
    ax.grid(True, alpha=0.25)
    ax.legend(loc="lower left")
    ax.annotate("Peak residual at 48s\n(longest recovery time)",
                xy=(48, -0.1919), xytext=(24, -0.16),
                fontsize=11, color="#FFCC00",
                arrowprops=dict(arrowstyle="->", color="#FFCC00", lw=1.5),
                ha="center")
    plt.tight_layout()
    out2 = OUT_DIR / "14_v2_dose_response_residual.png"
    plt.savefig(out2)
    plt.close(fig)
    print(f"Saved: {out2}")

    # -------- Chart 15: Within-window trajectories --------
    fig, ax = plt.subplots(figsize=(9, 6))
    for ice, _fn, color in RUNS:
        rows = data[ice]
        ts = [(int(r["sample"]) - 1) * 2 for r in rows]  # approx time in seconds within window
        toi = [float(r["toi_cal"]) for r in rows]
        ax.plot(ts, toi, marker="o", linewidth=2.0, markersize=10,
                color=color, markeredgecolor="#FFFFFF", markeredgewidth=1.0,
                label=f"{ice}s ice")
    ax.axhline(-0.21, color="#888888", linestyle=":", linewidth=1.2, alpha=0.6,
               label="Historical baseline")
    ax.set_xlabel("Time within capture window (s, after ice removal + transfer)")
    ax.set_ylabel("TOI_cal")
    ax.set_title("Within-Window TOI_cal Trajectories — All 5 V2 Runs")
    ax.invert_yaxis()
    ax.grid(True, alpha=0.25)
    ax.legend(loc="upper right", ncol=2)
    plt.tight_layout()
    out3 = OUT_DIR / "15_v2_trajectories.png"
    plt.savefig(out3)
    plt.close(fig)
    print(f"Saved: {out3}")

    # -------- Chart 16: NIRS vs PT100 overlay --------
    # Load PT100 fingertip data
    pt_temps = {}
    try:
        pt = openpyxl.load_workbook(PT100_SRC, data_only=True)
        ws = pt["Cooling Data"]
        for r in range(2, ws.max_row + 1):
            t = ws.cell(row=r, column=1).value
            v = ws.cell(row=r, column=2).value
            if t is not None and v is not None:
                pt_temps[int(t)] = float(v)
        # Linear interpolation at exact test durations
        def interp(t_target):
            if t_target in pt_temps:
                return pt_temps[t_target]
            times = sorted(pt_temps.keys())
            for i in range(len(times) - 1):
                t0, t1 = times[i], times[i+1]
                if t0 <= t_target <= t1:
                    v0, v1 = pt_temps[t0], pt_temps[t1]
                    return v0 + (v1 - v0) * (t_target - t0) / (t1 - t0)
            return None
        temps_at_runs = [interp(t) for t in ice_durations]
    except Exception as e:
        print(f"PT100 load error: {e}")
        temps_at_runs = [None]*5

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 9), sharex=True)
    # Top: PT100 cooling curve at test points
    times_sorted = sorted(pt_temps.keys())
    temps_sorted = [pt_temps[t] for t in times_sorted]
    valid = [(t, v) for t, v in zip(times_sorted, temps_sorted) if t <= 60]
    ax1.plot([x[0] for x in valid], [x[1] for x in valid], "-",
             color="#5AC8FA", linewidth=2, alpha=0.7,
             label="PT100 fingertip cooling")
    # Mark our test points on the curve
    for d, t in zip(ice_durations, temps_at_runs):
        if t is not None:
            ax1.plot(d, t, marker="o", markersize=14, color="#FF3B30",
                     markeredgecolor="#FFFFFF", markeredgewidth=1.5, zorder=5)
            ax1.annotate(f"{d}s\n{t:.1f}°C", (d, t), xytext=(8, -18),
                        textcoords="offset points", fontsize=10, color="#FFFFFF",
                        bbox=dict(boxstyle="round,pad=0.25", fc="#FF3B30", ec="none", alpha=0.85))
    ax1.set_ylabel("Skin temperature (°C)")
    ax1.set_title("Reference: Fingertip Skin Temperature During Cold Application")
    ax1.grid(True, alpha=0.25)
    ax1.legend(loc="upper right")

    # Bottom: NIRS cold peak at same ice durations
    ax2.plot(ice_durations, peaks, marker="o", markersize=14,
             linewidth=2.5, color="#FF3B30",
             markerfacecolor="#FF3B30", markeredgecolor="#FFFFFF",
             markeredgewidth=1.5, label="NIRS cold peak (sample 1)")
    for x, y in zip(ice_durations, peaks):
        ax2.annotate(f"{y:+.3f}", (x, y), xytext=(0, 14),
                    textcoords="offset points", ha="center",
                    fontsize=10, color="#FFFFFF",
                    bbox=dict(boxstyle="round,pad=0.25", fc="#FF3B30", ec="none", alpha=0.85))
    ax2.axhline(-0.21, color="#888888", linestyle=":", linewidth=1.2, alpha=0.6,
                label="Historical baseline (−0.21)")
    ax2.invert_yaxis()
    ax2.set_xlabel("Ice Duration (seconds)")
    ax2.set_ylabel("TOI_cal (less negative = more cold)")
    ax2.set_title("NIRS Cold-Peak Response at Same Ice Durations")
    ax2.set_xticks(ice_durations)
    ax2.grid(True, alpha=0.25)
    ax2.legend(loc="upper left")

    plt.tight_layout()
    out4 = OUT_DIR / "16_v2_nir_vs_pt100.png"
    plt.savefig(out4)
    plt.close(fig)
    print(f"Saved: {out4}")


if __name__ == "__main__":
    main()
