"""
Build comprehensive Excel for V2 fresh-ice cold therapy NIRS experiment.

Sheets:
  1. Summary       — dose-response table + key metrics
  2-6. Run 12/24/36/48/60s — raw per-sample data
  7. PT100 Ref     — fingertip cooling curve from prior experiment
  8. Predictions   — initial predictions vs actual
"""
import os
import csv
import math
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(r"c:/Users/corni/OneDrive - Nanyang Technological University/ntu_rf/M266_EmbeddedLinux")
V2DIR = BASE / "nir_intensity_20260527" / "v2_fresh_ice"
OUT = V2DIR / "v2_nir_intensity_dataset.xlsx"
PT100_SRC = BASE / "cold_therapy_showcase" / "12_skin_cooling_8sites.xlsx"

RUNS = [
    (12, "run1v2_12s.csv"),
    (24, "run2v2_24s.csv"),
    (36, "run3v2_36s.csv"),
    (48, "run4v2_48s.csv"),
    (60, "run5v2_60s.csv"),
]

THIN = Side(border_style="thin", color="888888")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="2D5A88", end_color="2D5A88", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HIGHLIGHT_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")


def parse_csv(p):
    rows = []
    with open(p, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(",")
            if parts[0] == "sample":
                header = parts
                continue
            row = {h: parts[i] for i, h in enumerate(header)}
            rows.append(row)
    return rows


def main():
    wb = openpyxl.Workbook()

    # ----- 1) Summary sheet -----
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "V2 Fresh-Ice NIRS Dose-Response — Fingertip"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")

    ws["A2"] = "Date: 2026-05-27 SGT"
    ws["A3"] = "Site: Middle finger fingertip"
    ws["A4"] = "Ice: fresh from freezer"
    ws["A5"] = "Capture: 5 samples × 2s = 10s post-ice"

    # Dose-response table (sample 1 = cold peak, sample 5 = stable point)
    headers = ["Run", "Ice duration (s)", "R_610 (S1)", "Cold peak TOI_cal (S1)",
               "Sample 5 TOI_cal (S5)", "Within-window Δ (S1→S5)",
               "Δ vs common baseline*"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BOX

    # Collect per-run data
    run_data = {}
    for ice_sec, fname in RUNS:
        rows = parse_csv(V2DIR / fname)
        run_data[ice_sec] = rows

    # Common baseline = mean of stable points of 12s/24s runs (least confounded by cold)
    s5_12 = float(run_data[12][-1]["toi_cal"])
    s5_24 = float(run_data[24][-1]["toi_cal"])
    common_baseline = (s5_12 + s5_24) / 2.0

    for r_idx, (ice_sec, fname) in enumerate(RUNS, start=8):
        rows = run_data[ice_sec]
        s1 = rows[0]
        s5 = rows[-1]
        s1_toi = float(s1["toi_cal"])
        s5_toi = float(s5["toi_cal"])
        within_delta = s1_toi - s5_toi
        common_delta = s1_toi - common_baseline

        ws.cell(row=r_idx, column=1, value=f"Run {r_idx-7}v2").border = BOX
        ws.cell(row=r_idx, column=2, value=ice_sec).border = BOX
        ws.cell(row=r_idx, column=3, value=int(s1["R_610"])).border = BOX
        c_peak = ws.cell(row=r_idx, column=4, value=round(s1_toi, 4))
        c_peak.border = BOX
        ws.cell(row=r_idx, column=5, value=round(s5_toi, 4)).border = BOX
        ws.cell(row=r_idx, column=6, value=round(within_delta, 4)).border = BOX
        ws.cell(row=r_idx, column=7, value=round(common_delta, 4)).border = BOX

    ws.cell(row=14, column=1, value="* Common baseline = mean of stable points of 12s+24s runs (least cold-confounded)")
    ws.cell(row=14, column=1).font = Font(italic=True, color="666666")
    ws.merge_cells("A14:G14")

    # Key findings text
    ws["A16"] = "Key findings:"
    ws["A16"].font = Font(bold=True)
    findings = [
        "1. Cold peak (sample 1) shows clear monotonic dose-response from 12s → 36s",
        "2. Saturation onset at 36-48s (peak plateau around −0.16)",
        "3. 60s shows small additional response (peak −0.146), consistent with saturating sigmoid",
        "4. Sample 5 (residual) peaks at 48s (most residual cold; reactive hyperemia not yet started)",
        "5. NOTE: Original 60s run (Cold peak −0.10) was REJECTED due to contact pressure artifact",
        "   (R_610 = 13710, well above v2 normal range 11000-12000). See run5v2_60s_REJECTED.csv.",
        "   Reproducibility check (R_610 = 11774) gave −0.146 — the canonical value used here.",
    ]
    for i, ftext in enumerate(findings, start=17):
        ws.cell(row=i, column=1, value=ftext)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

    for col_letter, w in [("A", 12), ("B", 16), ("C", 12), ("D", 22), ("E", 22), ("F", 22), ("G", 22)]:
        ws.column_dimensions[col_letter].width = w

    # ----- 2-6) Per-run sheets -----
    for ice_sec, fname in RUNS:
        sheet_name = f"Run {ice_sec}s"
        sh = wb.create_sheet(sheet_name)
        sh["A1"] = f"Run {ice_sec}v2 — {ice_sec}s ice, fresh pack, 10s capture"
        sh["A1"].font = Font(bold=True)
        sh.merge_cells("A1:K1")
        rows = run_data[ice_sec]
        cols = ["sample", "R_610", "S_680", "T_730", "U_760", "V_810", "W_860",
                "chip_temp", "toi_raw", "toi_cal", "sto2_dpf"]
        for c, h in enumerate(cols, start=1):
            cell = sh.cell(row=3, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BOX
        for r_idx, row in enumerate(rows, start=4):
            for c, h in enumerate(cols, start=1):
                v = row[h]
                if h in ("sample", "R_610", "S_680", "T_730", "U_760", "V_810", "W_860", "chip_temp"):
                    val = int(v)
                else:
                    val = float(v)
                cell = sh.cell(row=r_idx, column=c, value=val)
                cell.border = BOX
                # Highlight sample 1 (cold peak) and last sample (stable)
                if r_idx == 4 or r_idx == 4 + len(rows) - 1:
                    if h == "toi_cal":
                        cell.fill = HIGHLIGHT_FILL
        for col in "ABCDEFGHIJK":
            sh.column_dimensions[col].width = 12

    # ----- 7) PT100 Reference sheet -----
    try:
        pt = openpyxl.load_workbook(PT100_SRC, data_only=True)
        pt_ws = pt["Cooling Data"]
        ref_sh = wb.create_sheet("PT100 Ref (Fingertip)")
        ref_sh["A1"] = "Fingertip PT100 reference (from cold_therapy_showcase/12_skin_cooling_8sites.xlsx)"
        ref_sh["A1"].font = Font(bold=True)
        ref_sh.merge_cells("A1:C1")
        ref_sh["A2"] = "Note: anchor-fit model curve (not raw recorded data; anchors: tolerance=60s, endpoint~6°C)"
        ref_sh["A2"].font = Font(italic=True, color="666666")
        ref_sh.merge_cells("A2:C2")

        ref_sh.cell(row=4, column=1, value="Time (s)").font = HEADER_FONT
        ref_sh.cell(row=4, column=1).fill = HEADER_FILL
        ref_sh.cell(row=4, column=2, value="Skin temp (°C)").font = HEADER_FONT
        ref_sh.cell(row=4, column=2).fill = HEADER_FILL
        ref_sh.cell(row=4, column=3, value="Note").font = HEADER_FONT
        ref_sh.cell(row=4, column=3).fill = HEADER_FILL

        # Read fingertip column from PT100 source (column B in source)
        r = 5
        for src_row in range(2, pt_ws.max_row + 1):
            t = pt_ws.cell(row=src_row, column=1).value
            v = pt_ws.cell(row=src_row, column=2).value
            if t is None or v is None:
                continue
            ref_sh.cell(row=r, column=1, value=t).border = BOX
            ref_sh.cell(row=r, column=2, value=v).border = BOX
            # Mark our test points
            note = ""
            if t in (10, 25, 35, 50, 60):
                if t == 10:
                    note = "≈ 12s test (Cold peak −0.20)"
                elif t == 25:
                    note = "≈ 24s test (Cold peak −0.17)"
                elif t == 35:
                    note = "≈ 36s test (Cold peak −0.16)"
                elif t == 50:
                    note = "≈ 48s test (Cold peak −0.17)"
                elif t == 60:
                    note = "= 60s test (Cold peak −0.146, saturating)"
            ref_sh.cell(row=r, column=3, value=note).border = BOX
            if note:
                ref_sh.cell(row=r, column=3).fill = HIGHLIGHT_FILL
            r += 1

        ref_sh.column_dimensions["A"].width = 12
        ref_sh.column_dimensions["B"].width = 16
        ref_sh.column_dimensions["C"].width = 40
    except Exception as e:
        print(f"Warning: PT100 sheet skipped: {e}")

    # ----- 8) Predictions sheet -----
    pred = wb.create_sheet("Predictions vs Actual")
    pred["A1"] = "Pre-experiment predictions vs measured cold peak (sample 1)"
    pred["A1"].font = Font(bold=True)
    pred.merge_cells("A1:E1")

    cols = ["Ice (s)", "Predicted (linear)", "Predicted (saturating)", "Actual Cold peak", "Match"]
    for c, h in enumerate(cols, start=1):
        cell = pred.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BOX
    pred_data = [
        (12, -0.197, -0.20, -0.2019, "✓ both match"),
        (24, -0.174, -0.18, -0.1735, "✓ both match"),
        (36, -0.150, -0.16, -0.1636, "✓ saturating"),
        (48, -0.130, -0.159, -0.1663, "✓ saturating"),
        (60, -0.115, -0.156, -0.1458, "≈ saturating (slight extra)"),
    ]
    for i, row in enumerate(pred_data, start=4):
        for c, v in enumerate(row, start=1):
            cell = pred.cell(row=i, column=c, value=v)
            cell.border = BOX
            if row[0] == 60:
                cell.fill = HIGHLIGHT_FILL
    for col in "ABCDE":
        pred.column_dimensions[col].width = 22

    # Save
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
