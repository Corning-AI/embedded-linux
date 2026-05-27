"""
Multi-site NIRS dose-response analysis.

Fingertip: real data measured today (V2 fresh ice).
Other sites: projection based on PT100 cooling timing + fingertip TOI(temp) relationship.

Outputs (all in internal_only/):
  - multi_site_nir_dataset.xlsx  (4 sheets)
  - multi_site_pct_vs_temp.png   (% tolerance → skin temp, 8 sites)
  - multi_site_pct_vs_toi.png    (% tolerance → TOI projected, 8 sites)
  - multi_site_sensitivity.png   (TOI swing at 100% tolerance per site)
"""
import math
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

plt.rcParams.update({
    "font.size": 12, "axes.titlesize": 14, "axes.labelsize": 12,
    "legend.fontsize": 10,
    "figure.facecolor": "#0E0E14", "axes.facecolor": "#0E0E14",
    "axes.edgecolor": "#888888", "axes.labelcolor": "#DDDDDD",
    "xtick.color": "#DDDDDD", "ytick.color": "#DDDDDD",
    "text.color": "#FFFFFF", "axes.titlecolor": "#FFFFFF",
    "grid.color": "#333344", "savefig.facecolor": "#0E0E14",
    "savefig.edgecolor": "#0E0E14", "savefig.dpi": 150,
})

BASE = Path(r"c:/Users/corni/OneDrive - Nanyang Technological University/ntu_rf/M266_EmbeddedLinux")
OUT_DIR = BASE / "internal_only"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Site params: T_base, T_min, tau (s), tolerance (s), color
SITES = [
    ("指尖 Fingertip", 30.0, 5.0, 20, 60, "#FF3B30"),
    ("手背 Hand", 30.0, 10.0, 30, 90, "#FF9500"),
    ("膝盖 Knee", 29.0, 8.0, 35, 135, "#FFCC00"),
    ("前臂 Forearm", 33.0, 12.0, 50, 150, "#34C759"),
    ("肩膀 Shoulder", 33.0, 15.0, 80, 165, "#5AC8FA"),
    ("小腿 Calf", 31.0, 8.0, 70, 165, "#5856D6"),
    ("大腿 Thigh", 32.0, 10.0, 90, 210, "#AF52DE"),
    ("肚子 Belly", 33.7, 18.0, 110, 270, "#007AFF"),
]
BELLY_PLATEAU = 19.5

TOLERANCE_PCTS = [20, 40, 60, 80, 100]

# Today's fingertip dose-response (REAL DATA)
# (% tolerance, time, achieved temp from PT100 interp, TOI_cal cold peak)
FINGERTIP_REAL = [
    (20,  12, 18.7, -0.2019),
    (40,  24, 12.5, -0.1735),
    (60,  36,  9.1, -0.1636),
    (80,  48,  7.3, -0.1663),
    (100, 60,  6.4, -0.1458),
]
# Baseline: warm fingertip ~30°C → TOI ≈ -0.21
FP_TOI_CURVE_TEMPS = [30.0, 18.7, 12.5, 9.1, 7.3, 6.4]
FP_TOI_CURVE_VALUES = [-0.21, -0.2019, -0.1735, -0.1636, -0.1663, -0.1458]


def project_toi(temp):
    """Project TOI_cal at given skin temp by linear interp of fingertip TOI(temp) curve.
    Cap at warm baseline for temp > 30°C and at coldest measurement for temp < 6.4°C."""
    pts = sorted(zip(FP_TOI_CURVE_TEMPS, FP_TOI_CURVE_VALUES), reverse=True)
    if temp >= pts[0][0]:
        return pts[0][1]
    if temp <= pts[-1][0]:
        return pts[-1][1]
    for i in range(len(pts) - 1):
        t1, v1 = pts[i]
        t2, v2 = pts[i + 1]
        if t2 <= temp <= t1:
            frac = (t1 - temp) / (t1 - t2)
            return v1 + (v2 - v1) * frac
    return pts[-1][1]


def temp_at_time(t, T_base, T_min, tau):
    return T_min + (T_base - T_min) * math.exp(-t / tau)


# Compute table: per site × per %tolerance
table = {}
for name, T_base, T_min, tau, tol, color in SITES:
    table[name] = []
    for pct in TOLERANCE_PCTS:
        t = pct / 100.0 * tol
        temp = temp_at_time(t, T_base, T_min, tau)
        if "Belly" in name:
            temp = max(temp, BELLY_PLATEAU)
        toi = project_toi(temp)
        table[name].append({
            "pct": pct, "time": t, "temp": round(temp, 2),
            "toi_projected": round(toi, 4),
        })

# === Build Excel ===
xlsx_path = OUT_DIR / "multi_site_nir_dataset.xlsx"
wb = openpyxl.Workbook()

THIN = Side(border_style="thin", color="888888")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFILL = PatternFill(start_color="2D5A88", end_color="2D5A88", fill_type="solid")
HFONT = Font(bold=True, color="FFFFFF")
ROW_HL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

# Sheet 1: Site parameters
ws = wb.active
ws.title = "Site Parameters"
ws["A1"] = "Multi-site cooling parameters & projection inputs"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:G1")
ws["A2"] = "Fingertip: measured today (V2 fresh ice). Other sites: physics-based projection."
ws["A2"].font = Font(italic=True, color="888888")
ws.merge_cells("A2:G2")
headers = ["部位", "T_base (°C)", "T_min (°C)", "τ (s)", "Tolerance (s)",
           "vs Fingertip", "Notes"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = HFONT; cell.fill = HFILL; cell.border = BOX
for r, (name, T_base, T_min, tau, tol, _) in enumerate(SITES, start=5):
    vs_finger = round(tol / 60.0, 2)
    note = ""
    if "Fingertip" in name:
        note = "REAL data (today)"
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = ROW_HL
    elif "Belly" in name:
        note = "plateau correction at 19.5°C"
    ws.cell(row=r, column=1, value=name).border = BOX
    ws.cell(row=r, column=2, value=T_base).border = BOX
    ws.cell(row=r, column=3, value=T_min).border = BOX
    ws.cell(row=r, column=4, value=tau).border = BOX
    ws.cell(row=r, column=5, value=tol).border = BOX
    ws.cell(row=r, column=6, value=f"{vs_finger}×").border = BOX
    ws.cell(row=r, column=7, value=note).border = BOX
for col, w in zip("ABCDEFG", [22, 12, 12, 10, 14, 14, 30]):
    ws.column_dimensions[col].width = w

# Sheet 2: Skin temp at each %tolerance, per site
ws2 = wb.create_sheet("Temp at %tolerance")
ws2["A1"] = "Projected skin temperature at each %tolerance (from PT100 cooling formula)"
ws2["A1"].font = Font(bold=True, size=12)
ws2.merge_cells("A1:G1")
ws2.cell(row=3, column=1, value="部位").font = HFONT
ws2.cell(row=3, column=1).fill = HFILL
for c, pct in enumerate(TOLERANCE_PCTS, start=2):
    cell = ws2.cell(row=3, column=c, value=f"{pct}%")
    cell.font = HFONT; cell.fill = HFILL; cell.border = BOX
for c in range(1, 7):
    ws2.cell(row=3, column=c).border = BOX
for r, (name, *_rest) in enumerate(SITES, start=4):
    ws2.cell(row=r, column=1, value=name).border = BOX
    for c, pct in enumerate(TOLERANCE_PCTS, start=2):
        v = table[name][c - 2]["temp"]
        ws2.cell(row=r, column=c, value=v).border = BOX
    if "Fingertip" in name:
        for c in range(1, 7):
            ws2.cell(row=r, column=c).fill = ROW_HL
for col, w in zip("ABCDEF", [22, 10, 10, 10, 10, 10]):
    ws2.column_dimensions[col].width = w

# Sheet 3: Projected TOI per site per %tolerance
ws3 = wb.create_sheet("Projected TOI")
ws3["A1"] = "Projected TOI_cal at each %tolerance (fingertip REAL, others projected from PT100 temp)"
ws3["A1"].font = Font(bold=True, size=12)
ws3.merge_cells("A1:G1")
ws3.cell(row=3, column=1, value="部位").font = HFONT
ws3.cell(row=3, column=1).fill = HFILL
for c, pct in enumerate(TOLERANCE_PCTS, start=2):
    cell = ws3.cell(row=3, column=c, value=f"{pct}%")
    cell.font = HFONT; cell.fill = HFILL; cell.border = BOX
for c in range(1, 7):
    ws3.cell(row=3, column=c).border = BOX
for r, (name, *_rest) in enumerate(SITES, start=4):
    ws3.cell(row=r, column=1, value=name).border = BOX
    for c, pct in enumerate(TOLERANCE_PCTS, start=2):
        v = table[name][c - 2]["toi_projected"]
        ws3.cell(row=r, column=c, value=v).border = BOX
    if "Fingertip" in name:
        for c in range(1, 7):
            ws3.cell(row=r, column=c).fill = ROW_HL
for col, w in zip("ABCDEF", [22, 10, 10, 10, 10, 10]):
    ws3.column_dimensions[col].width = w

# Sheet 4: Today's fingertip real data
ws4 = wb.create_sheet("Fingertip Real Data")
ws4["A1"] = "Today's V2 fingertip dose-response (REAL measurements)"
ws4["A1"].font = Font(bold=True, size=12)
ws4.merge_cells("A1:E1")
hd = ["% tolerance", "Time (s)", "PT100 temp (°C)", "Cold peak TOI_cal", "Note"]
for c, h in enumerate(hd, start=1):
    cell = ws4.cell(row=3, column=c, value=h)
    cell.font = HFONT; cell.fill = HFILL; cell.border = BOX
for r, (pct, t, temp, toi) in enumerate(FINGERTIP_REAL, start=4):
    ws4.cell(row=r, column=1, value=pct).border = BOX
    ws4.cell(row=r, column=2, value=t).border = BOX
    ws4.cell(row=r, column=3, value=temp).border = BOX
    ws4.cell(row=r, column=4, value=toi).border = BOX
    ws4.cell(row=r, column=5, value="real" if pct < 999 else "").border = BOX
for col, w in zip("ABCDE", [14, 12, 16, 18, 18]):
    ws4.column_dimensions[col].width = w

wb.save(xlsx_path)
print(f"Saved Excel: {xlsx_path}")

# === Charts ===

# Chart 1: % tolerance → skin temp
fig, ax = plt.subplots(figsize=(10, 6))
for name, T_base, T_min, tau, tol, color in SITES:
    pcts = [t["pct"] for t in table[name]]
    temps = [t["temp"] for t in table[name]]
    en_label = name.split(" ", 1)[1] if " " in name else name
    ax.plot(pcts, temps, marker="o", linewidth=2.2, markersize=10,
            color=color, markeredgecolor="#FFFFFF", markeredgewidth=1.0,
            label=en_label)
ax.set_xlabel("% of cold tolerance")
ax.set_ylabel("Skin temperature (°C)")
ax.set_title("Multi-site comparison — skin temperature")
ax.set_xticks(TOLERANCE_PCTS)
ax.grid(True, alpha=0.25)
ax.legend(loc="upper right", ncol=2, fontsize=9)
plt.tight_layout()
c1 = OUT_DIR / "multi_site_pct_vs_temp.png"
plt.savefig(c1)
plt.close(fig)
print(f"Saved: {c1}")

# Chart 2: % tolerance → projected TOI
fig, ax = plt.subplots(figsize=(10, 6))
for name, T_base, T_min, tau, tol, color in SITES:
    pcts = [t["pct"] for t in table[name]]
    tois = [t["toi_projected"] for t in table[name]]
    en_label = name.split(" ", 1)[1] if " " in name else name
    ax.plot(pcts, tois, marker="o", linewidth=2.0, markersize=9,
            color=color, markeredgecolor="#FFFFFF", markeredgewidth=1.0,
            label=en_label)
ax.set_xlabel("% of cold tolerance")
ax.set_ylabel("Cold peak TOI_cal")
ax.set_title("Multi-site comparison — NIRS cold peak")
ax.invert_yaxis()
ax.set_xticks(TOLERANCE_PCTS)
ax.grid(True, alpha=0.25)
ax.axhline(-0.21, color="#888888", linestyle=":", linewidth=1.2, alpha=0.6,
           label="Historical warm baseline (−0.21)")
ax.legend(loc="lower right", ncol=2, fontsize=8.5)
plt.tight_layout()
c2 = OUT_DIR / "multi_site_pct_vs_toi.png"
plt.savefig(c2)
plt.close(fig)
print(f"Saved: {c2}")

# Chart 3: TOI swing at 100%tolerance per site
fig, ax = plt.subplots(figsize=(10, 6))
sites_sorted = sorted(SITES, key=lambda s: table[s[0]][-1]["toi_projected"])
names = [s[0].split(" ", 1)[1] if " " in s[0] else s[0] for s in sites_sorted]
swings_abs = [abs(table[s[0]][-1]["toi_projected"] - (-0.21)) for s in sites_sorted]
colors = [s[5] for s in sites_sorted]
bars = ax.barh(names, swings_abs, color=colors, edgecolor="#FFFFFF", linewidth=1.0)
for bar, val in zip(bars, swings_abs):
    ax.text(val + 0.002, bar.get_y() + bar.get_height() / 2,
            f"Δ={val:+.3f}", va="center", fontsize=10, color="#FFFFFF")
ax.set_xlabel("|TOI_cal swing| at 100% tolerance (vs warm baseline −0.21)")
ax.set_title("Multi-site comparison — sensitivity at full tolerance")
ax.grid(True, alpha=0.25, axis="x")
plt.tight_layout()
c3 = OUT_DIR / "multi_site_sensitivity.png"
plt.savefig(c3)
plt.close(fig)
print(f"Saved: {c3}")

print("\nDone.")
